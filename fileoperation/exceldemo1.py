from openpyxl import load_workbook
import mysql.connector
import entity
from openpyxl.styles import Font


def fun1():
    """
    读取excel
    :return:
    """
    workbook = load_workbook(filename="用户表.xlsx")
    print(workbook.sheetnames)
    sheet = workbook["傻狗表"]
    print(sheet)
    print(sheet.dimensions)  # A1:F3
    cell = sheet['A1']
    print(cell.value)
    print(cell.row, cell.column, cell.coordinate)
    cells = sheet['A1:A2']
    for row in cells:
        for cell in row:
            print(cell.value)

    print('-'.ljust(40, '-'))
    # 遍历所有行的数据
    for row in sheet.rows:
        print(row)


def fun2():
    """
    从数据库读取数据存储到列表里
    :return:
    """
    db = mysql.connector.connect(host="localhost", user="root", password="17770", database="mydb")
    cursor = db.cursor()
    sql = "select * from user"
    cursor.execute(sql)
    result = cursor.fetchall()
    users = []
    for row in result:
        user = entity.User(*row)
        users.append(user)
    return users


def fun3():
    """
    将列表数据存储到excel
    :return:
    """
    workbook = load_workbook(filename="用户表.xlsx")
    sheet = workbook["Sheet1"]
    users = fun2()
    sheet.append(["id", "code", "age", "name", "height", "address"])

    # 插入一列
    sheet.insert_cols(2)

    for user in users:
        sheet.append([user.id, user.code, user.age, user.name, user.height, user.address])
    workbook.save(filename="用户表.xlsx")


def fun4():
    """
    删除数据
    :return:
    """
    workbook = load_workbook(filename="用户表.xlsx")
    sheet = workbook["Sheet1"]
    # 删除某一行
    sheet.delete_rows(2)
    sheet.delete_cols(2)
    workbook.save(filename="用户表.xlsx")


def fun5():
    """
    批量调整字体样式
    :return:
    """
    workbook = load_workbook(filename="用户表.xlsx")
    sheet = workbook["Sheet1"]
    # 获取第一行
    row1 = sheet[1]
    for cell in row1:
        cell.font = Font(name='新魏', size=20, bold=True, italic=True, color="FF0000")
    workbook.save(filename="用户表.xlsx")


if __name__ == '__main__':
    # fun1()
    # users = fun2()
    # fun3()
    # fun4()
    fun5()
